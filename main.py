from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
from openpyxl import Workbook
import time
import os

import gspread
from google.oauth2.service_account import Credentials
from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive

# ✅ 設定
SPREADSHEET_ID = "19c6yIGr5BiI7XwstYhUPptFGksPPXE4N1bEq5iFoPok"
DRIVE_FOLDER_ID = "1MjNzGR57vsLtjbBJAZl06BKqZALYjGUO"
SERVICE_ACCOUNT_FILE = "service_account.json"
CREDENTIALS_FILE = "credentials.json"

# ✅ 日付設定
TODAY = datetime.now()
SHEET_NAME = TODAY.strftime("%y%m%d")
OUTPUT_FILE = f"{SHEET_NAME}.xlsx"

# ✅ 相対時間→日時変換
def parse_relative_time(text):
    now = datetime.now()
    try:
        if "分前" in text:
            return now - timedelta(minutes=int(text.replace("分前", "").strip()))
        elif "時間前" in text:
            return now - timedelta(hours=int(text.replace("時間前", "").strip()))
        elif "日前" in text:
            return now - timedelta(days=int(text.replace("日前", "").strip()))
        elif "秒前" in text:
            return now - timedelta(seconds=10)
    except:
        pass
    return now

def format_datetime(dt):
    return dt.strftime("%y/%m/%d %H:%M")

# ✅ Google SheetsからURL取得
def get_urls_from_sheet():
    scope = ["https://www.googleapis.com/auth/spreadsheets"]
    creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=scope)
    client = gspread.authorize(creds)
    sheet = client.open_by_key(SPREADSHEET_ID).worksheet(SHEET_NAME)
    rows = sheet.get_all_values()
    return rows, [row[2] for row in rows[1:] if len(row) > 2 and row[2].startswith("http")]

# ✅ Yahooニュース本文取得
def get_news_pages(base_url, driver):
    page_contents = []
    page = 1
    last_content = ""

    while True:
        url = base_url if page == 1 else f"{base_url}?page={page}"
        driver.get(url)
        time.sleep(2)
        try:
            article = driver.find_element(By.TAG_NAME, "article")
            paragraphs = article.find_elements(By.TAG_NAME, "p")
            content = "\n".join([p.text for p in paragraphs if p.text.strip()]).strip()
            if not content or content == last_content:
                break
            page_contents.append(content)
            last_content = content
            page += 1
        except:
            break

    driver.get(base_url)
    time.sleep(1)
    try:
        title = driver.title.replace(" - Yahoo!ニュース", "")
    except:
        title = "タイトル取得失敗"
    return title, base_url, page_contents

# ✅ コメント取得
def get_comments_pages(base_url, driver):
    comments_data = []
    page = 1
    last_comments = ""
    article_id = base_url.rstrip("/").split("/")[-1]
    base_comment_url = f"https://news.yahoo.co.jp/articles/{article_id}/comments"

    while True:
        comment_url = base_comment_url if page == 1 else f"{base_comment_url}?page={page}"
        driver.get(comment_url)
        time.sleep(2)
        soup = BeautifulSoup(driver.page_source, 'html.parser')
        comment_elements = soup.find_all('article', class_='sc-169yn8p-3 loqvSW')
        page_comments = []

        for comment_article in comment_elements:
            comment_p = comment_article.find('p', class_='sc-169yn8p-10 hYFULX')
            comment_text = comment_p.text.strip() if comment_p else ''
            user_a = comment_article.find('a', class_='sc-169yn8p-7 gibKWW')
            user_name = user_a.text.strip() if user_a else ''
            time_a = comment_article.find('a', class_='sc-169yn8p-9 gzAPCy')
            raw_time = time_a.text.strip() if time_a else ''
            dt = parse_relative_time(raw_time)
            formatted_time = format_datetime(dt)
            page_comments.append((comment_text, formatted_time, user_name))

        joined = "\n".join([c[0] for c in page_comments])
        if not page_comments or joined == last_comments:
            break

        last_comments = joined
        comments_data.extend(page_comments)
        page += 1

    return comments_data if comments_data else [("コメントなし", "", "")]

# ✅ Driveへアップロード
def upload_to_drive(file_path, folder_id):
    gauth = GoogleAuth()
    gauth.LoadCredentialsFile(CREDENTIALS_FILE)
    if gauth.credentials is None:
        gauth.LocalWebserverAuth()
    elif gauth.access_token_expired:
        gauth.Refresh()
    else:
        gauth.Authorize()
    gauth.SaveCredentialsFile(CREDENTIALS_FILE)

    drive = GoogleDrive(gauth)
    file = drive.CreateFile({'title': os.path.basename(file_path),
                             'parents': [{'id': folder_id}]})
    file.SetContentFile(file_path)
    file.Upload()
    print(f"✅ Driveアップロード完了: {file_path}")

# ✅ メイン処理
def main():
    print(f"📅 開始日付: {SHEET_NAME}")
    rows, urls = get_urls_from_sheet()
    if not urls:
        print("❌ URLが取得できませんでした")
        return

    options = Options()
    options.add_argument("--lang=ja-JP")
    options.add_argument("--headless")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    driver = webdriver.Chrome(options=options)

    wb = Workbook()
    ws_input = wb.active
    ws_input.title = "input"
    for i, row in enumerate(rows, 1):
        for j, val in enumerate(row, 1):
            ws_input.cell(row=i, column=j, value=val)

    for idx, url in enumerate(urls, 1):
        print(f"\n▶ ({idx}/{len(urls)}) 処理中: {url}")
        ws = wb.create_sheet(title=str(idx))

        try:
            title, base_url, pages = get_news_pages(url, driver)
            ws.cell(row=1, column=1, value="タイトル")
            ws.cell(row=1, column=2, value=title)
            ws.cell(row=2, column=1, value="URL")
            ws.cell(row=2, column=2, value=base_url)

            for i, page_text in enumerate(pages[:15], 1):
                ws.cell(row=i + 2, column=1, value=page_text)

            for i in range(len(pages)+3, 18):
                ws.cell(row=i, column=1, value="")  # 空白で埋める

        except Exception as e:
            ws.cell(row=1, column=1, value="エラー")
            ws.cell(row=2, column=1, value=str(e))

        try:
            comments = get_comments_pages(url, driver)
            start_row = 20
            ws.cell(row=start_row - 1, column=1, value="コメント本文")
            ws.cell(row=start_row - 1, column=2, value="投稿日時")
            ws.cell(row=start_row - 1, column=3, value="ユーザー名")
            for i, (text, dt, user) in enumerate(comments, start=start_row):
                ws.cell(row=i, column=1, value=text)
                ws.cell(row=i, column=2, value=dt)
                ws.cell(row=i, column=3, value=user)

            comment_count = len(comments) if comments[0][0] != "コメントなし" else 0
            ws_input.cell(row=idx + 1, column=6, value=comment_count)

        except Exception as e:
            ws.cell(row=20, column=1, value="コメント取得失敗")
            ws.cell(row=20, column=2, value=str(e))
            ws_input.cell(row=idx + 1, column=6, value="取得失敗")

    driver.quit()
    wb.save(OUTPUT_FILE)
    print(f"✅ Excel保存完了: {OUTPUT_FILE}")

    upload_to_drive(OUTPUT_FILE, DRIVE_FOLDER_ID)

if __name__ == "__main__":
    main()
